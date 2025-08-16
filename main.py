from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from apscheduler.schedulers.background import BackgroundScheduler
import os, time, uuid, csv, logging

app = FastAPI()

# CORS for local dev + production
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://127.0.0.1:5500",  # Live Server
        "http://localhost:5500",
        "http://127.0.0.1:8000",
        "http://localhost:8000",
        "https://your-frontend-domain.com"  # change for deployment
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)

TMP_DIR = "tmp"
os.makedirs(TMP_DIR, exist_ok=True)

DELIMITER_MAP = {
    "comma": ",",
    "tab": "\t",
    "semicolon": ";",
    "pipe": "|",
}
DELIMITER_NAME_MAP = {v: k.capitalize() for k, v in DELIMITER_MAP.items()}
DELIMITER_NAME_MAP.update({
    ",": "Comma",
    "\t": "Tab",
    ";": "Semicolon",
    "|": "Pipe"
})

MAX_ROWS_PER_SHEET = 1048575  # Excel limit minus header

def clean_tmp_folder():
    now = time.time()
    for filename in os.listdir(TMP_DIR):
        path = os.path.join(TMP_DIR, filename)
        if os.path.getmtime(path) < now - 3600:
            os.remove(path)
            logging.info(f"Deleted old file: {filename}")

scheduler = BackgroundScheduler()
scheduler.add_job(clean_tmp_folder, 'interval', minutes=30)
scheduler.start()

@app.get("/")
async def root():
    return {"message": "Service is running!"}

@app.post("/convert")
async def convert_file(
    file: UploadFile = File(...),
    delimiter: str = Form("auto"),
    encoding: str = Form("utf-8"),
):
    temp_upload_path = None
    try:
        if delimiter != "auto" and delimiter not in DELIMITER_MAP:
            raise ValueError("Invalid delimiter selected")
        if encoding not in ["utf-8", "latin-1", "utf-16"]:
            raise ValueError("Invalid encoding selected")

        content = await file.read()
        temp_upload_path = os.path.join(TMP_DIR, f"{uuid.uuid4().hex}_{file.filename}")
        with open(temp_upload_path, "wb") as f:
            f.write(content)

        with open(temp_upload_path, "r", encoding=encoding, errors="replace") as f:
            sample = f.read(4096)
        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            detected_delim = dialect.delimiter
            has_header = sniffer.has_header(sample)
        except Exception:
            detected_delim = ","
            has_header = True

        delim = DELIMITER_MAP.get(delimiter, detected_delim) if delimiter != "auto" else detected_delim

        wb = Workbook()
        sheet_num = 1
        sheet = wb.active
        sheet.title = f"Sheet_{sheet_num}"
        sheet_row_count = 0
        error_rows = []

        with open(temp_upload_path, "r", encoding=encoding, errors="replace") as f:
            reader = csv.reader(f, delimiter=delim)
            first_row = next(reader, None)
            if not first_row:
                raise ValueError("Empty file")

            headers = [str(cell).strip() for cell in first_row] if has_header else [f"Column_{i+1}" for i in range(len(first_row))]
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            sheet_row_count = 1

            if not has_header:
                sheet.append(first_row)
                sheet_row_count += 1

            for row_num, row in enumerate(reader, start=2):
                try:
                    if len(row) != len(headers):
                        raise ValueError("Incorrect number of columns")
                    if sheet_row_count >= MAX_ROWS_PER_SHEET:
                        sheet_num += 1
                        sheet = wb.create_sheet(f"Sheet_{sheet_num}")
                        sheet.append(headers)
                        for cell in sheet[1]:
                            cell.font = Font(bold=True)
                        sheet_row_count = 1
                    sheet.append(row)
                    sheet_row_count += 1
                except Exception as e:
                    error_rows.append((row_num, ",".join(row), str(e)))

        for s in wb.worksheets:
            for col in range(1, s.max_column + 1):
                letter = get_column_letter(col)
                max_len = min(max(len(str(cell.value or "")) for cell in s[letter]) + 2, 50)
                s.column_dimensions[letter].width = max_len
            s.freeze_panes = "A2"

        if error_rows:
            error_sheet = wb.create_sheet("Errors")
            error_sheet.append(["Row Number", "Raw Row", "Error"])
            for err in error_rows:
                error_sheet.append(err)

        temp_excel_path = os.path.join(TMP_DIR, f"{uuid.uuid4().hex}.xlsx")
        wb.save(temp_excel_path)

        os.remove(temp_upload_path)
        temp_upload_path = None

        detected_name = DELIMITER_NAME_MAP.get(detected_delim, "Custom")
        return {
            "download_url": f"/download/{os.path.basename(temp_excel_path)}",
            "detected_delimiter": detected_name,
            "file_size": os.path.getsize(temp_excel_path)
        }

    except Exception as e:
        if temp_upload_path and os.path.exists(temp_upload_path):
            os.remove(temp_upload_path)
        return JSONResponse(status_code=400, content={"error": str(e)})

@app.get("/download/{file_name}")
async def download_file(file_name: str, background_tasks: BackgroundTasks):
    path = os.path.join(TMP_DIR, file_name)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")

    background_tasks.add_task(os.remove, path)

    return FileResponse(
        path,
        filename="converted.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

